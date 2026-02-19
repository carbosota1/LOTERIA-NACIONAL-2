from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
from datetime import datetime, date as dt_date, timedelta
import os

from openpyxl import load_workbook

@dataclass
class Row:
    fecha: str
    sorteo: str
    primero: str
    segundo: str
    tercero: str

COLS = ["fecha", "sorteo", "primero", "segundo", "tercero"]

def _z2(x) -> str:
    s = str(x).strip()
    if s == "" or s.lower() == "none":
        return ""
    digits = "".join(ch for ch in s if ch.isdigit())
    if digits == "":
        return ""
    return digits.zfill(2)

def _date_to_ymd(x) -> str:
    if isinstance(x, dt_date):
        return x.strftime("%Y-%m-%d")
    s = str(x).strip()
    if " " in s:
        s = s.split(" ")[0].strip()
    datetime.strptime(s, "%Y-%m-%d")
    return s

def _make_key(fecha: str, sorteo: str) -> Tuple[str, str]:
    return (fecha, sorteo.strip())

def _ymd_to_date(s: str) -> dt_date:
    return datetime.strptime(s, "%Y-%m-%d").date()

def read_history_xlsx(path: str, sheet_name: str) -> List[Row]:
    if not os.path.exists(path):
        raise FileNotFoundError(f"No existe el history XLSX: {path}")

    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet_name}' en {path}. Hojas: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
    idx = {name: header.index(name) for name in COLS if name in header}
    missing = [c for c in COLS if c not in idx]
    if missing:
        raise ValueError(f"Faltan columnas en XLSX: {missing}. Header detectado: {header}")

    out: List[Row] = []
    for r in rows[1:]:
        if r is None:
            continue
        fecha_raw = r[idx["fecha"]]
        sorteo_raw = r[idx["sorteo"]]
        if fecha_raw is None or sorteo_raw is None:
            continue
        try:
            fecha = _date_to_ymd(fecha_raw)
        except Exception:
            continue
        sorteo = str(sorteo_raw).strip()
        primero = _z2(r[idx["primero"]])
        segundo = _z2(r[idx["segundo"]])
        tercero = _z2(r[idx["tercero"]])
        if not fecha or not sorteo:
            continue
        # si falta algún número, ignoramos (fila incompleta)
        if primero == "" or segundo == "" or tercero == "":
            continue
        out.append(Row(fecha, sorteo, primero, segundo, tercero))

    # dedupe por (fecha,sorteo) conservando el último
    m: Dict[Tuple[str, str], Row] = {}
    for rr in out:
        m[_make_key(rr.fecha, rr.sorteo)] = rr
    return list(m.values())

def write_full_history_xlsx(path: str, sheet_name: str, rows: List[Row]) -> None:
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet_name}' en {path}.")

    ws = wb[sheet_name]

    # obtener header existente
    existing = list(ws.iter_rows(values_only=True))
    if not existing:
        raise ValueError("La hoja 'history' está vacía (no hay header).")

    header = [str(x).strip().lower() if x is not None else "" for x in existing[0]]
    idx = {name: header.index(name) for name in COLS if name in header}
    missing = [c for c in COLS if c not in idx]
    if missing:
        raise ValueError(f"Faltan columnas en XLSX: {missing}. Header detectado: {header}")

    # limpiar todo menos header
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # ordenar
    rows_sorted = sorted(rows, key=lambda r: (r.fecha, r.sorteo))

    # escribir
    for rr in rows_sorted:
        ws.append([None] * len(header))  # fila vacía con largo correcto
        rownum = ws.max_row

        ws.cell(row=rownum, column=idx["fecha"] + 1, value=rr.fecha)
        ws.cell(row=rownum, column=idx["sorteo"] + 1, value=rr.sorteo)

        c1 = ws.cell(row=rownum, column=idx["primero"] + 1, value=_z2(rr.primero))
        c2 = ws.cell(row=rownum, column=idx["segundo"] + 1, value=_z2(rr.segundo))
        c3 = ws.cell(row=rownum, column=idx["tercero"] + 1, value=_z2(rr.tercero))

        # forzar texto para preservar 00/07
        c1.number_format = "@"
        c2.number_format = "@"
        c3.number_format = "@"

    wb.save(path)

def upsert_rows_xlsx(path: str, sheet_name: str, new_rows: List[Row]) -> int:
    if not new_rows:
        return 0

    existing = read_history_xlsx(path, sheet_name)
    m = {(r.fecha, r.sorteo): r for r in existing}
    changed = 0

    for nr in new_rows:
        key = (nr.fecha, nr.sorteo)
        prev = m.get(key)
        if prev is None or (prev.primero, prev.segundo, prev.tercero) != (nr.primero, nr.segundo, nr.tercero):
            m[key] = Row(nr.fecha, nr.sorteo, _z2(nr.primero), _z2(nr.segundo), _z2(nr.tercero))
            changed += 1

    if changed > 0:
        write_full_history_xlsx(path, sheet_name, list(m.values()))

    return changed

def sanitize_history_xlsx(path: str, sheet_name: str, today_ymd: str) -> Dict[str, int]:
    """
    Limpia el history:
      - elimina filas futuras (fecha > hoy)
      - elimina duplicados por (fecha,sorteo)
      - elimina feriado: mismo sorteo repite EXACTO (1ro,2do,3ro) en 2 días consecutivos => elimina el día más reciente
    Retorna contadores de removidos.
    """
    today = _ymd_to_date(today_ymd)
    rows = read_history_xlsx(path, sheet_name)

    before = len(rows)

    # 1) remover futuros
    rows_nf = []
    removed_future = 0
    for r in rows:
        try:
            d = _ymd_to_date(r.fecha)
        except Exception:
            continue
        if d > today:
            removed_future += 1
            continue
        rows_nf.append(r)

    # 2) organizar por sorteo y remover feriados (consecutivos iguales)
    by_draw: Dict[str, List[Row]] = {}
    for r in rows_nf:
        by_draw.setdefault(r.sorteo, []).append(r)

    removed_holiday = 0
    cleaned: List[Row] = []

    for sorteo, lst in by_draw.items():
        lst_sorted = sorted(lst, key=lambda x: x.fecha)
        keep: List[Row] = []
        prev: Optional[Row] = None

        for cur in lst_sorted:
            if prev is None:
                keep.append(cur)
                prev = cur
                continue

            prev_date = _ymd_to_date(prev.fecha)
            cur_date = _ymd_to_date(cur.fecha)

            # si son consecutivos y los 3 números son idénticos => feriado / repetición fantasma => eliminar cur
            if (cur_date - prev_date).days == 1 and (cur.primero, cur.segundo, cur.tercero) == (prev.primero, prev.segundo, prev.tercero):
                removed_holiday += 1
                # NO actualizamos prev, porque mantenemos el prev real
                continue

            keep.append(cur)
            prev = cur

        cleaned.extend(keep)

    after = len(cleaned)

    # reescribir si cambió algo
    if after != before or removed_future > 0 or removed_holiday > 0:
        write_full_history_xlsx(path, sheet_name, cleaned)

    return {
        "before": before,
        "after": after,
        "removed_future": removed_future,
        "removed_holiday": removed_holiday,
        "removed_total": before - after + removed_future  # nota: before/after ya incluye dedupe; esto es indicativo
    }

def latest_before(rows: List[Row], sorteo: str, fecha_ymd: str) -> Optional[Row]:
    """
    Devuelve la fila más reciente del sorteo con fecha < fecha_ymd
    """
    target = _ymd_to_date(fecha_ymd)
    candidates = []
    for r in rows:
        if r.sorteo != sorteo:
            continue
        try:
            d = _ymd_to_date(r.fecha)
        except Exception:
            continue
        if d < target:
            candidates.append((d, r))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[-1][1]
